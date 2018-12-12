# coding=utf-8
import json
from urllib import request, parse

from openpyxl import Workbook
from openpyxl import load_workbook


def opens():
    wb = load_workbook(r"C:\Users\ness\Downloads\关键词4374.xlsx")
    ws = wb.active
    data = []
    for row in ws.rows:
        list = []
        for cell in row:
            list.append(str(cell.value))
        data.append(list)

    return data


def transfer(word, langauge):
    headers = {
        "Host": "fanyi.youdao.com",
        "Origin": "http://fanyi.youdao.com",
        "Referer": "http://fanyi.youdao.com",
        "User-Agent": "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36",
        "Cookie": "UM_distinctid=15d1bb7488b99-007bc3345a9dbb-474a0721-1fa400-15d1bb7488c5fb; JSESSIONID=aaaQ3r7qEJVL-gQ4vWB0v; SESSION_FROM_COOKIE=fanyiweb; OUTFOX_SEARCH_USER_ID=-1302043367@125.120.80.215; _ntes_nnid=e1833dfbbef8cd5f51dbc5016ddc34e8,1499409188679; OUTFOX_SEARCH_USER_ID_NCOO=1614799888.6369958; ___rl__test__cookies=1499410377257"

    }
    datas = {
        "i": word,
        "from": "zh - CHS",
        "to": langauge,
        "smartresult": "dict",
        "client": "fanyideskweb",
        # "salt": int(time.time()),
        "salt": "1499410377261",
        "sign": "e449ee788ebe3766347736207412f811",
        "doctype": "json",
        "version": "2.1",
        "keyfrom": "fanyi.web",
        "action": "Y_BY_CLICKBUTTON",
        "typoResult": "true"
    }

    req = request.Request(
        url="http://fanyi.youdao.com/translate_o?smartresult=dict&smartresult=rule&sessionFrom=https://www.baidu.com/link",
        headers=headers)
    res = request.urlopen(req, data=parse.urlencode(datas).encode('utf-8'))

    html = json.loads(res.read())
    html["translateResult"][0][0]["tgt"]



def saves(arr):
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active

    for line in arr:
        ws.append(line)
        print(line)

        # 保存文件
    wb.save("sample.xlsx")


languages={

    "英文":"en",



}

if __name__ == '__main__':
    transfer("吾爱APP ASO关键词", "ko")
